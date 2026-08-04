#!/usr/bin/env python3
"""将灯饰售后 Excel 文件转换为紧凑 JSON 格式，供 Web 分析系统使用。"""
import json, sys, os, time, re
from collections import Counter
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl，正在安装...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl


# ── 文件路径 ──
import glob as _glob

def _find_latest(desktop, pattern):
    files = _glob.glob(os.path.join(desktop, pattern))
    return max(files, key=os.path.getmtime) if files else None

DESKTOP = os.path.join(os.environ.get("USERPROFILE", ""), "Desktop")
VALID_SKU_PREFIXES = {"DJ", "CL", "PL", "CF", "WL", "CD", "TL", "FL"}

# 检测合并文件（含"售后明细"sheet）
MERGED_FILE = None
if len(sys.argv) > 1 and sys.argv[1].lower().endswith((".xlsx", ".xls")):
    _wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
    if "售后明细" in _wb.sheetnames:
        MERGED_FILE = sys.argv[1]
    _wb.close()

if MERGED_FILE:
    AFTERSALE_FILE = SALES_FILE = MERGED_FILE
    INPUT_MODE = "merged"
    print(f"[合并模式] 读取: {MERGED_FILE}")
else:
    AFTERSALE_FILE = (
        sys.argv[1] if len(sys.argv) > 1 and sys.argv[1].lower().endswith((".xlsx", ".xls")) else
        _find_latest(DESKTOP, "独立站灯饰售后工单*.xlsx")
    )
    SALES_FILE = _find_latest(DESKTOP, "独立站灯饰销量统计*.xlsx") if len(sys.argv) <= 1 else (
        sys.argv[2] if len(sys.argv) > 2 else _find_latest(DESKTOP, "独立站灯饰销量统计*.xlsx")
    )
    INPUT_MODE = "split"
    if not AFTERSALE_FILE or not os.path.exists(AFTERSALE_FILE):
        print("[X] 未找到售后工单文件，请在桌面放置 '独立站灯饰售后工单*.xlsx' 或拖入 xlsx")
        sys.exit(1)
    if not SALES_FILE or not os.path.exists(SALES_FILE):
        print("[X] 未找到销量统计文件，请在桌面放置 '独立站灯饰销量统计*.xlsx'")
        sys.exit(1)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")
COMPACT_FILE = os.path.join(OUTPUT_DIR, "after-sale-data-compact.json")
VERSION_FILE = os.path.join(OUTPUT_DIR, "version.json")


# ── 责任方分类 ──
def classify_responsibility(tag, reason):
    text = f"{tag or ''} {reason or ''}"
    if "越兴" in text:
        return "D-采购-越兴相关"
    for kw in ["停产", "晚发货", "下架不可售", "漏采"]:
        if kw in text:
            return "D-采购-越兴相关"
    for kw in ["发错货", "漏发", "发错", "错发", "少发", "多发"]:
        if kw in text:
            return "B-仓库-发货问题"
    for kw in ["物流", "运输", "快递", "破损", "丢件", "丢失", "海关", "转运", "第三方"]:
        if kw in text:
            return "F-物流-运输问题"
    for kw in ["品质", "质量", "灯不亮", "不亮", "外观", "烧坏", "烧毁", "安全", "隐患",
               "故障", "损坏", "缺陷", "不良", "开裂", "生锈", "掉漆", "刮花", "划痕", "工艺", "尺寸", "颜色"]:
        if kw in text:
            return "A-品控-品质问题"
    for kw in ["不喜欢", "不想要", "无理由", "下错", "下单错", "大小",
               "买错", "拍错", "重复", "不要了", "退货", "个人", "不满", "折扣",
               "不需要", "风格", "色差", "交期"]:
        if kw in text:
            return "C-客户-个人原因"
    for kw in ["描述", "说明", "图物", "不符", "安装", "网页", "listing", "页面", "参数"]:
        if kw in text:
            return "E-运营-信息问题"
    return "C-客户-个人原因"


# ── 工具函数 ──
def safe_float(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default

def safe_int(v, default=0):
    try: return int(float(v)) if v is not None else default
    except: return default

def parse_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, (int, float)):
        if val > 25569:
            base = datetime(1899, 12, 30)
            return (base + __import__('datetime').timedelta(days=int(val))).strftime("%Y-%m-%d %H:%M:%S")
    s = str(val).strip()
    return s[:19] if s else ""

def parse_year(val):
    if val is None:
        return 0
    if isinstance(val, datetime):
        return val.year
    s = str(val).strip()[:10]
    if s and s[:4].isdigit():
        return int(s[:4])
    return 0


# ── 读取 Excel ──
def read_excel(filepath, sheet_name=None):
    print(f"读取: {filepath}" + (f" [sheet={sheet_name}]" if sheet_name else ""))
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    print(f"  → {len(rows)} 行, 表头: {headers[:8]}...")
    return headers, rows

def find_col(headers, *aliases):
    for h in aliases:
        if h in headers:
            return headers.index(h)
    for i, h in enumerate(headers):
        for a in aliases:
            if a.lower() in h.lower():
                return i
    return -1


# ── 主流程 ──
def main():
    t_start = time.time()
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 1. 读取售后
    ar_sheet = "售后明细" if INPUT_MODE == "merged" else None
    ah, ar_rows = read_excel(AFTERSALE_FILE, ar_sheet)
    col = {h: i for i, h in enumerate(ah)}

    after_sale_records = []
    type_counter = Counter()
    reason_counter = Counter()
    tag_counter = Counter()
    resp_counter = Counter()
    sku_stats = {}
    yearly = {}

    for row in ar_rows:
        sku_val = str(row[find_col(ah, "Sku", "SKU")]).strip() if find_col(ah, "Sku", "SKU") >= 0 and row[find_col(ah, "Sku", "SKU")] else ""
        if not sku_val or sku_val[:2] not in VALID_SKU_PREFIXES:
            continue

        tid = str(row[find_col(ah, "售后工单单号", "ticket_id")]).strip() if find_col(ah, "售后工单单号", "ticket_id") >= 0 and row[find_col(ah, "售后工单单号", "ticket_id")] else ""
        after_type = str(row[find_col(ah, "售后类型")]).strip() if find_col(ah, "售后类型") >= 0 and row[find_col(ah, "售后类型")] else ""
        reason = str(row[find_col(ah, "售后原因")]).strip() if find_col(ah, "售后原因") >= 0 and row[find_col(ah, "售后原因")] else ""
        tag = str(row[find_col(ah, "标签")]).strip() if find_col(ah, "标签") >= 0 and row[find_col(ah, "标签")] else ""
        prod_name = str(row[find_col(ah, "产品名称")]).strip() if find_col(ah, "产品名称") >= 0 and row[find_col(ah, "产品名称")] else ""

        ct = parse_date(row[find_col(ah, "创建时间", "创建人/创建时间", "创建时间/创建人", "创建日期")]) if find_col(ah, "创建时间", "创建人/创建时间", "创建时间/创建人", "创建日期") >= 0 else ""
        ot = parse_date(row[find_col(ah, "订购时间", "下单时间", "订单时间")]) if find_col(ah, "订购时间", "下单时间", "订单时间") >= 0 else ""
        rs = str(row[find_col(ah, "退款状态")]).strip() if find_col(ah, "退款状态") >= 0 and row[find_col(ah, "退款状态")] else ""
        oa = safe_float(row[find_col(ah, "订单金额")]) if find_col(ah, "订单金额") >= 0 else 0
        ra = safe_float(row[find_col(ah, "退款金额")]) if find_col(ah, "退款金额") >= 0 else 0
        rq = safe_int(row[find_col(ah, "退货数量")], 1) if find_col(ah, "退货数量") >= 0 else 1

        year = parse_year(row[find_col(ah, "创建时间", "创建人/创建时间", "创建时间/创建人", "创建日期")]) if find_col(ah, "创建时间", "创建人/创建时间", "创建时间/创建人", "创建日期") >= 0 else 0
        if year < 2020:
            year = parse_year(row[find_col(ah, "订购时间", "下单时间", "订单时间")]) if find_col(ah, "订购时间", "下单时间", "订单时间") >= 0 else 0

        resp = classify_responsibility(tag, reason)

        type_counter[after_type] += 1
        reason_counter[reason] += 1
        tag_counter[tag] += 1
        resp_counter[resp] += 1

        if year > 0:
            if year not in yearly:
                yearly[year] = {"after_sale_count": 0, "sales_count": 0, "total_orders": 0, "total_return_qty": 0}
            yearly[year]["after_sale_count"] += 1
            yearly[year]["total_return_qty"] += rq

        if sku_val not in sku_stats:
            sku_stats[sku_val] = {"count": 0, "total_return_qty": 0, "total_refund": 0}
        sku_stats[sku_val]["count"] += 1
        sku_stats[sku_val]["total_return_qty"] += rq
        sku_stats[sku_val]["total_refund"] += ra

        after_sale_records.append({
            "tid": tid, "t": after_type, "p": prod_name, "sku": sku_val,
            "r": reason, "rs": rs, "oa": oa, "ra": ra, "rq": rq,
            "tg": tag, "ct": ct, "resp": resp, "y": year,
        })

    print(f"  售后记录: {len(after_sale_records)}")

    # 2. 读取销量
    sr_sheet = "销量明细" if INPUT_MODE == "merged" else None
    sh, sr_rows = read_excel(SALES_FILE, sr_sheet)

    sales_records = []
    total_sales_qty = 0
    total_orders = 0

    for row in sr_rows:
        sku_val = str(row[find_col(sh, "SKU")]).strip() if find_col(sh, "SKU") >= 0 and row[find_col(sh, "SKU")] else ""
        if not sku_val or sku_val[:2] not in VALID_SKU_PREFIXES:
            continue

        date_val = row[find_col(sh, "时间", "日期")]
        d_str = parse_date(date_val) if find_col(sh, "时间", "日期") >= 0 else ""
        year = parse_year(date_val) if find_col(sh, "时间", "日期") >= 0 else 0

        sq = safe_int(row[find_col(sh, "销量")]) if find_col(sh, "销量") >= 0 else 0
        oq = safe_int(row[find_col(sh, "订单量")], 1) if find_col(sh, "订单量") >= 0 else 1
        rv = safe_float(row[find_col(sh, "销售额")]) if find_col(sh, "销售额") >= 0 else 0
        pf = str(row[find_col(sh, "平台")]).strip() if find_col(sh, "平台") >= 0 and row[find_col(sh, "平台")] else ""
        pn = str(row[find_col(sh, "品名", "产品名称")]).strip() if find_col(sh, "品名", "产品名称") >= 0 and row[find_col(sh, "品名", "产品名称")] else ""

        total_sales_qty += sq
        total_orders += oq

        if year > 0:
            if year not in yearly:
                yearly[year] = {"after_sale_count": 0, "sales_count": 0, "total_orders": 0, "total_return_qty": 0}
            yearly[year]["sales_count"] += 1
            yearly[year]["total_orders"] += oq

        sales_records.append({
            "d": d_str, "sku": sku_val, "pn": pn, "sq": sq, "oq": oq,
            "rv": rv, "pf": pf, "y": year,
        })

    print(f"  销量记录: {len(sales_records)}")

    # 3. 统计
    all_skus = set(r["sku"] for r in after_sale_records) | set(r["sku"] for r in sales_records)

    # SKU TOP20
    sku_top20 = sorted(
        [{"sku": k, "product_name": "", "count": v["count"],
          "total_return_qty": v["total_return_qty"], "total_refund": round(v["total_refund"], 2)}
         for k, v in sku_stats.items()],
        key=lambda x: -x["total_return_qty"]
    )[:20]

    # 统计
    responsibilities = []
    for k, count in resp_counter.most_common():
        responsibilities.append({"responsibility": k, "count": count,
            "total_return_qty": 0, "total_refund": 0,
            "pct": round(count / len(after_sale_records) * 100, 1),
            "avg_refund": 0})

    after_sale_types = [{"after_sale_type": k, "count": v,
        "pct": round(v / len(after_sale_records) * 100, 1)} for k, v in type_counter.most_common()]
    reasons = [{"reason": k, "count": v,
        "pct": round(v / len(after_sale_records) * 100, 1)} for k, v in reason_counter.most_common()]
    tags = [{"tag": k, "count": v,
        "pct": round(v / len(after_sale_records) * 100, 1)} for k, v in tag_counter.most_common()]

    # 日期范围
    dates_after = sorted([r["ct"][:10] for r in after_sale_records if r["ct"]])
    dates_sales = sorted([r["d"][:10] for r in sales_records if r["d"]])

    # 4. 构建 compact JSON
    compact = {
        "m": {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_after_sale": len(after_sale_records),
            "total_sales": len(sales_records),
            "total_orders": total_orders,
            "date_range": {
                "after_sale_min": dates_after[0] if dates_after else "",
                "after_sale_max": dates_after[-1] if dates_after else "",
                "sales_min": dates_sales[0] if dates_sales else "",
                "sales_max": dates_sales[-1] if dates_sales else "",
            },
            "years": sorted(yearly.keys()),
        },
        "ys": {str(k): v for k, v in yearly.items()},
        "st": {
            "after_sale_types": after_sale_types,
            "reasons": reasons,
            "tags": tags,
            "responsibilities": responsibilities,
            "sku_top20": sku_top20,
        },
        "ar": after_sale_records,
        "sr": sales_records,
    }

    # 5. 写入
    with open(COMPACT_FILE, "w", encoding="utf-8") as f:
        json.dump(compact, f, ensure_ascii=False, separators=(",", ":"))
    file_size = os.path.getsize(COMPACT_FILE)
    print(f"\n写入: {COMPACT_FILE} ({file_size / 1024 / 1024:.1f} MB)")

    version_data = {
        "version": "1.0.0",
        "updated": compact["m"]["generated_at"],
        "data_checksum": str(file_size),
        "stats": {
            "after_sale_count": len(after_sale_records),
            "sales_count": len(sales_records),
        },
    }
    with open(VERSION_FILE, "w", encoding="utf-8") as f:
        json.dump(version_data, f, ensure_ascii=False, indent=2)

    print(f"✅ 转换完成，耗时 {time.time() - t_start:.1f}s")


if __name__ == "__main__":
    main()
